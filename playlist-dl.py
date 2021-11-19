#!/usr/bin/env python3

from __future__ import unicode_literals

import sys
from urllib.parse import urlparse, urlencode, parse_qs, urlunparse

import openpyxl
import youtube_dl


def clean_url(url):
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    if not query:
        return None
    new_query_params = {'v': query['v']}
    return urlunparse(parsed._replace(query=urlencode(new_query_params, True)))


def get_youtube_links(workbook_name, sheet_num, start_row, start_col):
    base_col = openpyxl.utils.column_index_from_string(start_col)
    workbook = openpyxl.load_workbook(workbook_name)
    worksheet = workbook[workbook.sheetnames[sheet_num]]
    for row in range(start_row, worksheet.max_row):
        ytb_link = worksheet.cell(column=base_col, row=row).value
        if ytb_link is not None:
            yield ytb_link


filename = sys.argv[1]
sheet_num = int(sys.argv[2])
cell = sys.argv[3]

links = get_youtube_links(filename, sheet_num, int(cell[1]), cell[0])

clean_links = [clean_url(link) for link in links]
clean_links = [link for link in clean_links if link]


class SilentLogger(object):
    def debug(self, msg):
        pass
    def warning(self, msg):
        pass
    def error(self, msg):
        print(msg)


def my_hook(d):
    if d['status'] == 'finished':
        print(f'downloaded song={d["filename"]}, now converting ...')

ydl_opts = {
    'rm_cachedir': True,
    'format': 'bestaudio/best',
    'postprocessors': [{
        'key': 'FFmpegExtractAudio',
        'preferredcodec': 'mp3',
        'preferredquality': '192',
    }],
    'logger': SilentLogger(),
    'progress_hooks': [my_hook],
    'outtmpl': '%(autonumber)02d-%(title)s.%(ext)s',
}

with youtube_dl.YoutubeDL(ydl_opts) as ydl:
    print(clean_links)
    try:
        ydl.download(clean_links)
    except Exception as e:
        print(e)
